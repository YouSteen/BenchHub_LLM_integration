import os
import pandas as pd
import win32com.client as win32
from openpyxl import load_workbook

from menu.options.send_emails.llm_integration.run import generate_llm_outputs
from menu.utils.config_manager import get_survey_path, get_llm_path

# === CONFIG ===
subject = "Your Development Plan – Personalized Suggestions"

EMAIL_TEMPLATE = """
<html>
<body>
<p>Hello <b>{name}</b>,</p>
<p>Thank you for completing the form! I’d like to share some valuable information to support your upskilling and certification journey.</p>
{generated_section}
<p>These resources will help you develop your skills in the fields you’re most passionate about. Feel free to explore the available materials, and don’t hesitate to reach out to Local Community Leads if you have any questions or need further support.</p>
<p style="margin-top: 20px;"><b>Romania Testing Technical Communities: Collaboration & Knowledge Hub</b></p>
</body>
</html>
"""


def send_email_outlook(to: str, cc: str, subject: str, html_body: str):
    outlook = win32.Dispatch("Outlook.Application")
    mail = outlook.CreateItem(0)
    mail.To = to
    mail.CC = cc
    mail.Subject = subject
    mail.HTMLBody = html_body
    mail.Send()


def send_all_emails():
    survey_path = get_survey_path()
    llm_model_path = get_llm_path()

    if not survey_path or not os.path.exists(survey_path):
        print("Survey file not found. Please configure the path in the options menu.")
        input("Press Enter to return to the menu...")
        return

    if not llm_model_path or not os.path.exists(llm_model_path):
        print("LLM model not found. Please configure the path in the options menu.")
        input("Press Enter to return to the menu...")
        return

    # Check LLM model size (must be > 4GB)
    if os.path.getsize(llm_model_path) < 4 * 1024 * 1024 * 1024:
        print("LLM model file is smaller than 4GB. Please select a valid model.")
        input("Press Enter to return to the menu...")
        return

    # Check for necessary columns in the survey file
    df = pd.read_excel(survey_path)
    required_columns = [
        "Id",
        "Start time",
        "Completion time",
        "Email",
        "Name",
        "Please choose the office location you are assigned to.\n",
        "Please provide your Career Coach's email address in the space below.\n",
        "Which areas of upskilling are you most interested in? (You can select multiple options)\n",
        "Are you currently engaged in any training, certifications, or testing activities?\n",
        "If yes, please specify the type of training or certification you're currently engaged with.\n",
        "Are there any specific topics or skills you would like to focus on in future training programs?\n",
        "Are you available to actively participate in testing communities and contribute to different activities that support their growth and development?\n",
        "Would you be open to sharing your expertise by leading training sessions within your field of knowledge?\n",
        "Would you be interested in joining Keystone MarketPlace Romania in the future?\nhttps://confluence.endava.com/spaces/CTO/pages/577044538/KMP+in+Romania",
        "Do you feel you have enough information or clarity regarding your next steps during bench period?\n",
        "Send Email",
    ]

    missing_columns = [col for col in required_columns if col not in df.columns]

    if missing_columns:
        print(
            f"The survey file is missing the following required columns: {', '.join(missing_columns)}"
        )
        input("Press Enter to return to the menu...")
        return

    wb = load_workbook(survey_path)
    ws = wb.active

    if ws is None:
        raise ValueError("No active worksheet found in the Excel file.")

    col_send = [c for c in df.columns if "send email" in c.lower()][0]
    col_email = [c for c in df.columns if "email" in c.lower()][0]

    responses = generate_llm_outputs(survey_path)

    for idx, entry in enumerate(responses):
        name = entry["name"]
        email = entry["email"]
        coach = entry["coach"]
        body = entry["llm_output"]

        html = EMAIL_TEMPLATE.format(name=name, generated_section=body)

        print(f"Sending to {name} <{email}> | CC: {coach}")
        send_email_outlook(to=email, cc=coach, subject=subject, html_body=html)

        matching_rows = df[df[col_email] == email]
        if not matching_rows.empty:
            excel_row = matching_rows.index[0]
            column_location = df.columns.get_loc(col_send)
            if isinstance(excel_row, int) and isinstance(column_location, int):
                ws.cell(row=excel_row + 2, column=column_location + 1).value = 1

    wb.save(survey_path)
    print("All emails sent and Excel updated.")


if __name__ == "__main__":
    send_all_emails()
